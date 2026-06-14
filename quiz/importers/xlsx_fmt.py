"""XLSX import/export with friendly column names, dropdowns and separate SA regex column."""
import io

from quiz import grading
from quiz.importers.base import (ParsedQuestion, correct_to_spec,
                                 parse_correct_spec, validate_question)

MAX_CHOICES = 6

# ── Display ↔ internal code mappings ──────────────────────────────────────────

TYPE_DISPLAY = {
    grading.MC: 'Multiple Choice',
    grading.MA: 'Multiple Answer',
    grading.TF: 'True/False',
    grading.SA: 'Short Answer',
}
_TYPE_PARSE = {}
for _code, _label in TYPE_DISPLAY.items():
    _TYPE_PARSE[_label.lower()] = _code
    _TYPE_PARSE[_code.lower()] = _code   # also accept old abbreviations

LEVEL_DISPLAY = {'easy': 'Easy', 'medium': 'Medium', 'hard': 'Hard'}
_LEVEL_PARSE = {v.lower(): k for k, v in LEVEL_DISPLAY.items()}
_LEVEL_PARSE.update({k: k for k in LEVEL_DISPLAY})  # accept lowercase too

MA_STRATEGY_DISPLAY = {
    grading.MA_ALL_OR_NOTHING:    'All or nothing',
    grading.MA_PARTIAL_CREDIT:    'Partial credit',
    grading.MA_RIGHT_MINUS_WRONG: 'Right minus wrong',
    grading.MA_CORRECT_ONLY:      'Correct only',
}
_MA_STRATEGY_PARSE = {v.lower(): k for k, v in MA_STRATEGY_DISPLAY.items()}
_MA_STRATEGY_PARSE.update({k: k for k in MA_STRATEGY_DISPLAY})

TRUTHY = ('yes', 'true', '1', 'x', 'có')

# ── Column layout ─────────────────────────────────────────────────────────────

# Internal keys (used for parsing)
# Interleaved: choice_N then choice_explanation_N
HEADERS = (
    'code',
    'type', 'title', 'content',
    'choice_1', 'choice_explanation_1',
    'choice_2', 'choice_explanation_2',
    'choice_3', 'choice_explanation_3',
    'choice_4', 'choice_explanation_4',
    'choice_5', 'choice_explanation_5',
    'choice_6', 'choice_explanation_6',
    'correct',
    'points', 'category', 'level',
    'explanation', 'shuffle', 'ma_strategy',
    'answer_display',
)

# Human-readable labels shown in row 1
HEADER_LABELS = (
    'Code',
    'Type', 'Title', 'Question',
    'Choice 1', 'Explanation 1',
    'Choice 2', 'Explanation 2',
    'Choice 3', 'Explanation 3',
    'Choice 4', 'Explanation 4',
    'Choice 5', 'Explanation 5',
    'Choice 6', 'Explanation 6',
    'Correct Answer',
    'Points', 'Category', 'Level',
    'Explanation', 'Shuffle Choices', 'MA Strategy',
    'Answer Display',
)


# Pre-compute column letters (A=0, B=1, ...)
def _col_letter(header_key):
    return chr(ord('A') + HEADERS.index(header_key))


HEADER_NOTES = {
    'Code': (
        'Unique short identifier for this question — lowercase letters and '
        'digits only, no spaces or underscores. Examples: py101q1, tfindent3'
    ),
    'Type': (
        'Question type — choose from the dropdown:\n'
        '  Multiple Choice  : single correct choice\n'
        '  Multiple Answer  : one or more correct choices\n'
        '  True/False       : boolean statement\n'
        '  Short Answer     : student types a text answer'
    ),
    'Correct Answer': (
        'MC  : choice number, e.g.  2\n'
        'MA  : comma-separated numbers, e.g.  1,3\n'
        'TF  : True  or  False\n'
        'SA  : regex patterns separated by  |  (each is tested with re.fullmatch)\n'
        '      Use (?i) prefix for case-insensitive:  (?i)python\n'
        '      Use | to offer alternatives:  3|three\n'
        '      Example:  (?i)def  |  (?i)define\n'
        '      Digits: \\d+   Any word: \\w+'
    ),
    'MA Strategy': (
        'How Multiple Answer questions are scored — choose from dropdown:\n'
        '  All or nothing    : full score only if every correct choice selected\n'
        '  Partial credit    : proportional score minus wrong-answer penalty\n'
        '  Right minus wrong : (correct − wrong) / total_correct\n'
        '  Correct only      : correct / total_correct, no wrong-answer penalty'
    ),
    'Shuffle Choices': 'Yes — randomise the choice order shown to each student.',
    'Answer Display': (
        'SA only — human-readable answer shown to students on the result page.\n'
        'Leave blank to fall back to showing the raw regex patterns.\n'
        'Example: if patterns are  (?i)paris , write  Paris  here.'
    ),
    'Level': 'Difficulty — choose from dropdown: Easy, Medium, Hard',
    'Question': (
        'Question body shown to students. Markdown is supported:\n'
        '  **bold**   _italic_   `inline code`   $math$\n'
        '  Fenced code blocks, tables, and lists are also supported.'
    ),
    'Choice 1': 'Choice text for option 1. Markdown is supported (**bold**, _italic_, `code`, etc.).',
    'Choice 2': 'Choice text for option 2. Markdown is supported (**bold**, _italic_, `code`, etc.).',
    'Choice 3': 'Choice text for option 3. Markdown is supported (**bold**, _italic_, `code`, etc.).',
    'Choice 4': 'Choice text for option 4. Markdown is supported (**bold**, _italic_, `code`, etc.).',
    'Choice 5': 'Choice text for option 5. Markdown is supported (**bold**, _italic_, `code`, etc.).',
    'Choice 6': 'Choice text for option 6. Markdown is supported (**bold**, _italic_, `code`, etc.).',
    'Explanation 1': (
        'Optional per-choice explanation for option 1, shown via a "Why?" toggle after submitting.\n'
        'Markdown is supported (**bold**, _italic_, `code`, etc.).'
    ),
    'Explanation 2': (
        'Optional per-choice explanation for option 2, shown via a "Why?" toggle after submitting.\n'
        'Markdown is supported (**bold**, _italic_, `code`, etc.).'
    ),
    'Explanation 3': (
        'Optional per-choice explanation for option 3, shown via a "Why?" toggle after submitting.\n'
        'Markdown is supported (**bold**, _italic_, `code`, etc.).'
    ),
    'Explanation 4': (
        'Optional per-choice explanation for option 4, shown via a "Why?" toggle after submitting.\n'
        'Markdown is supported (**bold**, _italic_, `code`, etc.).'
    ),
    'Explanation 5': (
        'Optional per-choice explanation for option 5, shown via a "Why?" toggle after submitting.\n'
        'Markdown is supported (**bold**, _italic_, `code`, etc.).'
    ),
    'Explanation 6': (
        'Optional per-choice explanation for option 6, shown via a "Why?" toggle after submitting.\n'
        'Markdown is supported (**bold**, _italic_, `code`, etc.).'
    ),
    'Explanation': (
        'Overall question explanation shown at the bottom of the result page.\n'
        'Markdown is supported: **bold**, _italic_, `code`, tables, lists, etc.'
    ),
}

# 23 columns: A through W
COL_WIDTHS = {
    'A': 16,                          # code
    'B': 18,  'C': 28,  'D': 46,     # type, title, content
    'E': 22,  'F': 32,               # choice_1, explanation_1
    'G': 22,  'H': 32,               # choice_2, explanation_2
    'I': 22,  'J': 32,               # choice_3, explanation_3
    'K': 22,  'L': 32,               # choice_4, explanation_4
    'M': 22,  'N': 32,               # choice_5, explanation_5
    'O': 22,  'P': 32,               # choice_6, explanation_6
    'Q': 32,                          # correct
    'R': 8,   'S': 18,  'T': 12,    # points, category, level
    'U': 40,  'V': 14,  'W': 18,    # explanation, shuffle, ma_strategy
    'X': 28,                          # answer_display
}

# ── Row helper: each row is exactly 22 values matching HEADERS ────────────────
# col order: type, title, content,
#            ch1, ex1, ch2, ex2, ch3, ex3, ch4, ex4, ch5, ex5, ch6, ex6,
#            correct, points, category, level, explanation, shuffle, ma_strategy


def _mc(code, title, content, choices_expls, correct_1based, points,
        category, level, explanation, shuffle=False):
    """choices_expls: list of (text, explanation) tuples."""
    row = [code, 'Multiple Choice', title, content]
    for i in range(MAX_CHOICES):
        if i < len(choices_expls):
            row += list(choices_expls[i])
        else:
            row += [None, None]
    row += [str(correct_1based), points, category, level,
            explanation, 'Yes' if shuffle else None, None, None]
    return tuple(row)


def _ma(code, title, content, choices_expls, correct_csv, points,
        category, level, explanation, strategy='Partial credit'):
    row = [code, 'Multiple Answer', title, content]
    for i in range(MAX_CHOICES):
        if i < len(choices_expls):
            row += list(choices_expls[i])
        else:
            row += [None, None]
    row += [correct_csv, points, category, level, explanation, 'Yes', strategy, None]
    return tuple(row)


def _tf(code, title, content, correct_bool, points, category, level, explanation):
    row = [code, 'True/False', title, content] + [None] * (MAX_CHOICES * 2)
    row += ['True' if correct_bool else 'False', points, category,
            level, explanation, None, None, None]
    return tuple(row)


def _sa(code, title, content, patterns, points, category, level, explanation,
        answer_display=''):
    row = [code, 'Short Answer', title, content] + [None] * (MAX_CHOICES * 2)
    row += [' | '.join(patterns), points, category, level,
            explanation, None, None, answer_display or None]
    return tuple(row)


# ── Example questions: Python programming (20 questions, all types) ───────────

EXAMPLE_ROWS = (
    # ── Multiple Choice: Easy (3) ─────────────────────────────────────────────
    _mc(
        code='pymcfloordiv',
        title='Floor division operator result',
        content='What is the result of `7 // 2` in Python?',
        choices_expls=[
            ('3.5', 'This is regular division (`7 / 2`), not floor division.'),
            ('3',   '✓ Floor division (`//`) divides and truncates toward negative infinity, giving 3.'),
            ('4',   'Rounding up would give 4, but floor division always rounds toward negative infinity.'),
            ('2',   'This would be the remainder (`7 % 2`), not the quotient.'),
        ],
        correct_1based=2, points=1, category='Python Basics', level='Easy',
        explanation='The `//` operator performs floor division, which divides two numbers and truncates the result toward negative infinity. `7 // 2` equals 3 because 7 divided by 2 is 3.5, and the floor of 3.5 is 3.',
        shuffle=True,
    ),

    _mc(
        code='pymclistappend',
        title='list.append() vs list.extend()',
        content='What does `my_list.append([4, 5])` do when `my_list = [1, 2, 3]`?',
        choices_expls=[
            ('Adds 4 and 5 as separate elements, making `my_list` have 5 elements',
             'This describes `extend()`, not `append()`. `extend()` iterates and adds each item separately.'),
            ('Raises a `TypeError` because you cannot append a list to a list',
             'Python allows appending any object, including lists, to a list.'),
            ('Adds `[4, 5]` as a single element, making `my_list` have 4 elements',
             '✓ `append()` always adds exactly one object to the end; the list `[4, 5]` becomes a nested element.'),
            ('Concatenates the lists and returns a new list `[1, 2, 3, 4, 5]`',
             '`append()` modifies in place and returns `None`; `+` concatenation creates a new list.'),
        ],
        correct_1based=3, points=1, category='Python Iterables', level='Easy',
        explanation="`list.append(x)` adds `x` as a single element to the end of the list, regardless of `x`'s type. To merge all elements of an iterable into the list, use `extend()` instead.",
        shuffle=True,
    ),

    _mc(
        code='pymcstrtype',
        title='Type of a string literal',
        content='What is the type of `"42"` in Python?',
        choices_expls=[
            ('`int`',   '`int` represents integer numbers. `"42"` has quotes, making it text, not a number.'),
            ('`float`', '`float` represents decimal numbers. Quotes make this a string regardless of content.'),
            ('`str`',   '✓ Any value enclosed in single or double quotes is a `str` (string) object in Python.'),
            ('`bytes`', '`bytes` literals use a `b` prefix, e.g. `b"42"`. Plain quotes produce a `str`.'),
        ],
        correct_1based=3, points=1, category='Python Data Types', level='Easy',
        explanation='In Python, any value enclosed in single quotes (`\'...\'`) or double quotes (`"..."`) is a `str` object. The fact that the content looks numeric is irrelevant — the quotes determine the type.',
        shuffle=False,
    ),

    # ── Multiple Choice: Medium (3) ───────────────────────────────────────────
    _mc(
        code='pymcdefaultarg',
        title='Mutable default argument pitfall',
        content=(
            'What is printed by the following code?\n\n'
            '```python\n'
            'def add_item(item, lst=[]):\n'
            '    lst.append(item)\n'
            '    return lst\n\n'
            'print(add_item(1))\n'
            'print(add_item(2))\n'
            '```'
        ),
        choices_expls=[
            ('[1]\n[2]',
             'This would be correct if `lst` were re-created on each call, but mutable defaults are shared across calls.'),
            ('[1]\n[1, 2]',
             '✓ The default list `[]` is created once when the function is defined and reused on every call that omits `lst`, so mutations accumulate.'),
            ('[1]\n[2, 1]',
             'Items are appended in call order; there is no reversal here.'),
            ('A `TypeError` is raised on the second call',
             'The code is syntactically and semantically valid Python; no error is raised.'),
        ],
        correct_1based=2, points=2, category='Python Functions', level='Medium',
        explanation='Mutable default arguments (like `[]` or `{}`) are evaluated **once** when the `def` statement runs, not on every call. All calls that rely on the default share the same object, so mutations persist between calls. The fix is to use `None` as the default and create a new list inside the function body.',
        shuffle=False,
    ),

    _mc(
        code='pymcisinstance',
        title='isinstance() with inheritance',
        content=(
            'Given:\n\n'
            '```python\n'
            'class Animal: pass\n'
            'class Dog(Animal): pass\n'
            'd = Dog()\n'
            '```\n\n'
            'Which expression evaluates to `True`?'
        ),
        choices_expls=[
            ('`type(d) is Animal`',
             '`type()` returns the exact class; `d` is exactly a `Dog`, not an `Animal`, so this is `False`.'),
            ('`type(d) is Dog`',
             'This is `True`, but `isinstance(d, Animal)` is more informative — it covers the full inheritance chain.'),
            ('`isinstance(d, Animal)`',
             '✓ `isinstance()` checks the full MRO; since `Dog` inherits from `Animal`, `d` is an instance of both.'),
            ('`isinstance(d, str)`',
             '`d` is a `Dog`; it has no relationship to `str`, so this is `False`.'),
        ],
        correct_1based=3, points=2, category='Python OOP', level='Medium',
        explanation='`isinstance(obj, cls)` returns `True` if `obj` is an instance of `cls` *or any subclass thereof*. Because `Dog` inherits from `Animal`, `isinstance(d, Animal)` is `True`. By contrast, `type(d) is Animal` returns `False` because `type()` checks for exact class identity.',
        shuffle=False,
    ),

    _mc(
        code='pymcgenexpr',
        title='Generator expression vs comprehension',
        content='Which of the following creates a **generator** (lazy iterator) rather than building the full sequence in memory?',
        choices_expls=[
            ('`[x**2 for x in range(10)]`',
             'Square brackets produce a **list comprehension**, which evaluates all elements immediately.'),
            ('`{x**2 for x in range(10)}`',
             'Curly braces without a colon produce a **set comprehension**, not a generator.'),
            ('`(x**2 for x in range(10))`',
             '✓ Parentheses (without an outer function call) produce a **generator expression** — a lazy iterator that yields values one at a time.'),
            ('`tuple(x**2 for x in range(10))`',
             '`tuple()` *consumes* the generator and materialises all values into memory; the result is not a generator.'),
        ],
        correct_1based=3, points=2, category='Python Iterables', level='Medium',
        explanation='A generator expression uses parentheses `(expr for var in iterable)` and returns a generator object that yields values lazily. List `[...]`, set `{...}`, and dict `{k:v ...}` comprehensions all materialise their results immediately.',
        shuffle=False,
    ),

    # ── Multiple Choice: Hard (3) ─────────────────────────────────────────────
    _mc(
        code='pymcclosureval',
        title='Closure late-binding behaviour',
        content=(
            'What does the following code print?\n\n'
            '```python\n'
            'funcs = [lambda: i for i in range(3)]\n'
            'print([f() for f in funcs])\n'
            '```'
        ),
        choices_expls=[
            ('[0, 1, 2]',
             'This would be correct if each lambda captured the *value* of `i` at creation time, but Python closures capture the *variable*, not its value.'),
            ('[2, 2, 2]',
             '✓ All three lambdas close over the same variable `i`. By the time they are called, the loop has finished and `i` is 2, so every lambda returns 2.'),
            ('[0, 0, 0]',
             'Closures do not freeze the initial value of the variable; they reference it dynamically.'),
            ('A `NameError` because `i` is out of scope',
             'The loop variable `i` persists in the enclosing scope after the loop ends; no `NameError` occurs.'),
        ],
        correct_1based=2, points=3, category='Python Functions', level='Hard',
        explanation='Python closures capture variables by *reference*, not by value. All three lambdas share the same `i` variable from the enclosing scope. After the `for` loop completes, `i` equals 2, so every call to `f()` returns 2. The classic fix is to use a default argument: `lambda i=i: i`.',
        shuffle=False,
    ),

    _mc(
        code='pymcmroorder',
        title='MRO method resolution order',
        content=(
            'Given the following class hierarchy, what does `C().greet()` print?\n\n'
            '```python\n'
            'class A:\n'
            '    def greet(self): return "A"\n\n'
            'class B(A):\n'
            '    def greet(self): return "B"\n\n'
            'class C(B, A):\n'
            '    pass\n'
            '```'
        ),
        choices_expls=[
            ('"A"', '`A.greet` is further up the MRO than `B.greet`; Python searches left-to-right so `B` is checked before `A`.'),
            ('"B"', "✓ Python's C3 linearisation MRO for `C` is `[C, B, A, object]`. The first class that defines `greet` is `B`."),
            ('"C"', '`C` does not define `greet`, so it cannot return `"C"`.'),
            ('An `AttributeError`', '`greet` is inherited from `B`; no `AttributeError` is raised.'),
        ],
        correct_1based=2, points=3, category='Python OOP', level='Hard',
        explanation="Python uses the C3 linearisation algorithm to compute the MRO. For `class C(B, A)`, the MRO is `C → B → A → object`. Python walks this list left-to-right and calls the first `greet` it finds, which belongs to `B`. You can inspect any class's MRO with `C.__mro__`.",
        shuffle=False,
    ),

    _mc(
        code='pymcdescriptor',
        title='Property vs instance __dict__ priority',
        content=(
            'What does the following print?\n\n'
            '```python\n'
            'class Circle:\n'
            '    def __init__(self, r):\n'
            '        self.__dict__["radius"] = r\n\n'
            '    @property\n'
            '    def radius(self):\n'
            '        return self.__dict__["radius"] * 2\n\n'
            'c = Circle(5)\n'
            'print(c.radius)\n'
            '```'
        ),
        choices_expls=[
            ('5',
             'If the instance `__dict__` shadowed the property, you would get 5, but data descriptors take priority.'),
            ('10',
             '✓ A `property` is a data descriptor; it takes precedence over same-named keys in `__dict__`. The getter runs and returns `5 * 2 = 10`.'),
            ('An `AttributeError`',
             'Storing directly into `__dict__` bypasses the descriptor protocol for *setting*, so no error occurs during `__init__`.'),
            ('An infinite recursion error',
             'The getter reads from `__dict__` directly with `self.__dict__["radius"]`, avoiding recursive property access.'),
        ],
        correct_1based=2, points=3, category='Python OOP', level='Hard',
        explanation='In Python\'s descriptor protocol, **data descriptors** (objects defining both `__get__` and `__set__`, like `property`) take priority over instance `__dict__` entries with the same name. So `c.radius` always invokes the getter even though `__dict__["radius"]` was set directly. The getter returns `5 * 2 = 10`.',
        shuffle=False,
    ),

    # ── Multiple Answer (4, one of each MA strategy) ──────────────────────────
    _ma(
        code='pymabuiltins',
        title='Built-in functions that return iterators',
        content='Which of the following built-in functions return a **lazy iterator** (not a list) in Python 3?',
        choices_expls=[
            ('`map()`',    '✓ In Python 3, `map()` returns a map object (lazy iterator), not a list.'),
            ('`filter()`', '✓ In Python 3, `filter()` returns a filter object (lazy iterator), not a list.'),
            ('`sorted()`', '`sorted()` always returns a new `list`; it is never lazy.'),
            ('`zip()`',    '✓ `zip()` returns a zip object (lazy iterator) that yields tuples on demand.'),
            ('`list()`',   '`list()` consumes an iterable and returns a fully materialised list immediately.'),
        ],
        correct_csv='1,2,4', points=2, category='Python Iterables', level='Medium',
        explanation='In Python 3, `map()`, `filter()`, and `zip()` all return lazy iterator objects that produce values one at a time. `sorted()` and `list()` always materialise their results into a list.',
        strategy='Partial credit',
    ),

    _ma(
        code='pymaimutset',
        title='Immutable types in Python',
        content='Which of the following Python types are **immutable** (cannot be changed after creation)?',
        choices_expls=[
            ('`tuple`',     '✓ Tuples cannot be modified after creation; any "modification" produces a new tuple.'),
            ('`list`',      '`list` is mutable — elements can be added, removed, or reassigned in place.'),
            ('`frozenset`', '✓ `frozenset` is the immutable counterpart of `set`; it supports no add/remove operations.'),
            ('`str`',       '✓ Strings are immutable; operations like `s.upper()` always return a *new* string.'),
            ('`dict`',      '`dict` is mutable; keys and values can be added, removed, or updated.'),
        ],
        correct_csv='1,3,4', points=2, category='Python Data Types', level='Medium',
        explanation='Immutable types cannot be modified after creation: `tuple`, `frozenset`, `str`, `int`, `float`, and `bytes`. Mutable types (`list`, `dict`, `set`) support in-place modification.',
        strategy='All or nothing',
    ),

    _ma(
        code='pymaexcbase',
        title='Base classes of ValueError',
        content='Which of the following are **direct or indirect base classes** of `ValueError`?',
        choices_expls=[
            ('`Exception`',    '✓ `ValueError` inherits from `Exception`, the base for all non-system-exiting exceptions.'),
            ('`BaseException`', '✓ `BaseException` is the root of the entire exception hierarchy; every exception is a subclass.'),
            ('`RuntimeError`', '`RuntimeError` and `ValueError` are sibling subclasses of `Exception`; neither inherits from the other.'),
            ('`LookupError`',  '`LookupError` is the base of `KeyError`/`IndexError`, not `ValueError`.'),
            ('`object`',       '✓ In Python, every class ultimately inherits from `object`, including all exceptions.'),
        ],
        correct_csv='1,2,5', points=3, category='Python Exceptions', level='Hard',
        explanation='The MRO for `ValueError` is: `ValueError → Exception → BaseException → object`. `RuntimeError` and `LookupError` are *siblings* (both inherit from `Exception`) but are not ancestors of `ValueError`.',
        strategy='Right minus wrong',
    ),

    _ma(
        code='pymadunder',
        title='Dunder methods for two-way + support',
        content='Which dunder methods should a class implement to support the `+` operator **both ways** (i.e. `a + b` and `b + a` when `b` is a different type)?',
        choices_expls=[
            ('`__add__`',  '✓ `__add__` handles left-hand-side `a + b`; Python calls `a.__add__(b)` first.'),
            ('`__radd__`', "✓ `__radd__` is the reflected addition, called when the left operand's `__add__` returns `NotImplemented`."),
            ('`__iadd__`', '`__iadd__` handles *in-place* addition (`a += b`), not plain `+` with a foreign type on the left.'),
            ('`__sum__`',  'There is no `__sum__` dunder method in Python; `sum()` uses `__add__` internally.'),
            ('`__pos__`',  '`__pos__` is the unary plus operator (`+a`), not binary addition.'),
        ],
        correct_csv='1,2', points=3, category='Python OOP', level='Hard',
        explanation='To support `a + b` where `a` is your type, implement `__add__`. To also support `b + a` where `b` is a built-in or foreign type, implement `__radd__`. Python first tries `b.__add__(a)`; if that returns `NotImplemented`, it falls back to `a.__radd__(b)`. `__iadd__` handles `+=` separately.',
        strategy='Correct only',
    ),

    # ── True / False (4) ──────────────────────────────────────────────────────
    _tf(
        code='pytfindent',
        title='Indentation is syntactically significant',
        content='In Python, indentation is **syntactically significant** — using inconsistent indentation within a block raises an `IndentationError`.',
        correct_bool=True, points=1, category='Python Basics', level='Easy',
        explanation='Unlike many languages that use braces `{}` to delimit code blocks, Python uses indentation as part of its syntax. Mixing tabs and spaces, or using inconsistent indent levels within a block, raises an `IndentationError` at parse time.',
    ),

    _tf(
        code='pytfnonelies',
        title='None is falsy in boolean context',
        content='`None` evaluates to `False` in a boolean context (e.g. inside an `if` statement).',
        correct_bool=True, points=1, category='Python Basics', level='Easy',
        explanation='Python treats several values as falsy: `None`, `False`, `0`, `0.0`, empty sequences (`""`, `[]`, `()`), and empty mappings (`{}`). `None` is not equal to `False`, but `bool(None)` returns `False`, so it behaves as falsy in conditionals.',
    ),

    _tf(
        code='pytftupmut',
        title='Tuple containing a list is fully immutable',
        content='A tuple that contains a list is fully immutable — neither the tuple nor any of its elements can be changed.',
        correct_bool=False, points=2, category='Python Data Types', level='Medium',
        explanation="A tuple's *references* are immutable — you cannot replace, add, or remove elements from the tuple itself. However, if a tuple element is a mutable object (like a list), that object's internal state *can* be changed. For example, `t = ([1, 2],); t[0].append(3)` works fine.",
    ),

    _tf(
        code='pytfglobalkw',
        title='Assignment inside function creates local variable',
        content='Without the `global` keyword, assigning to a variable inside a function **always** creates a new local variable, even if a global variable with the same name exists.',
        correct_bool=True, points=2, category='Python Functions', level='Medium',
        explanation="Python's scoping rule (LEGB) means that any assignment inside a function makes that name local to the function by default. Without `global x`, writing `x = value` inside a function shadows the global `x` rather than modifying it. The `global` keyword is needed to rebind a global variable from within a function.",
    ),

    # ── Short Answer (3) ──────────────────────────────────────────────────────
    _sa(
        code='pysarangelen',
        title='Built-in for sequence length',
        content=(
            'What built-in function returns the number of elements in a `range` object '
            'without converting it to a list? Type the function name only (e.g. `foo`).'
        ),
        patterns=[r'(?i)len'],
        points=1, category='Python Iterables', level='Easy',
        explanation='`len()` works directly on `range` objects in Python 3 — it computes the count in O(1) using the start, stop, and step values without iterating. For example, `len(range(0, 10, 2))` returns `5`.',
        answer_display='len',
    ),

    _sa(
        code='pysawalrus',
        title='Walrus operator symbol',
        content=(
            'Python 3.8 introduced the **assignment expression** operator, nicknamed the "walrus operator". '
            'What is its symbol? Type just the operator (e.g. `+=`).'
        ),
        patterns=[r':='],
        points=2, category='Python Basics', level='Medium',
        explanation='The walrus operator `:=` (PEP 572, Python 3.8) assigns a value to a variable as part of an expression. Common use: `while chunk := f.read(8192)` or `if m := re.search(pattern, text)`. It avoids repeating an expression just to capture its result.',
        answer_display=':=',
    ),

    _sa(
        code='pysaslotsdunder',
        title='Dunder for memory-efficient attributes',
        content=(
            'What dunder attribute can be defined in a Python class to restrict instance attributes '
            'to a fixed set, eliminating the per-instance `__dict__` and reducing memory usage? '
            'Type the attribute name including underscores (e.g. `__foo__`).'
        ),
        patterns=[r'__slots__'],
        points=3, category='Python OOP', level='Hard',
        explanation='Defining `__slots__ = ("x", "y")` in a class tells Python to allocate a fixed-size array for those attributes instead of a per-instance `__dict__`. This reduces memory usage significantly for classes with many instances and also slightly speeds up attribute access.',
        answer_display='__slots__',
    ),
)


# ── SA helper ────────────────────────────────────────────────────────────────

def _sa_correct_to_cell(correct_answers):
    """SA correct_answers (list of regex strings or legacy dicts) → cell string."""
    parts = []
    for a in (correct_answers or []):
        parts.append(a.get('text', '') if isinstance(a, dict) else str(a))
    return ' | '.join(parts) or None


# ── Public API ────────────────────────────────────────────────────────────────

def parse(file_obj):
    """file-like → list[ParsedQuestion]."""
    from openpyxl import load_workbook
    try:
        sheet = load_workbook(file_obj, read_only=True,
                              data_only=True).active
    except Exception as exc:
        broken = ParsedQuestion(row=0)
        broken.errors.append('Cannot read XLSX file: %s' % exc)
        return [broken]

    questions = []
    for row_num, cells in enumerate(
            sheet.iter_rows(min_row=2, values_only=True), start=2):
        cells = list(cells) + [None] * (len(HEADERS) - len(cells))
        if all(c in (None, '') for c in cells):
            continue
        data = dict(zip(HEADERS, cells))
        q = ParsedQuestion(row=row_num)
        questions.append(q)

        q.code = str(data.get('code') or '').strip().lower()
        raw_type = str(data['type'] or '').strip()
        q.type = _TYPE_PARSE.get(raw_type.lower(), raw_type.upper())

        q.title = str(data['title'] or '').strip()
        q.content = str(data['content'] or '')
        raw_choices = [
            str(data['choice_%d' % i]).strip()
            for i in range(1, MAX_CHOICES + 1)
            if data['choice_%d' % i] not in (None, '')
        ]
        all_explanations = [
            str(data.get('choice_explanation_%d' % i) or '').strip()
            for i in range(1, MAX_CHOICES + 1)
        ]
        all_explanations = all_explanations[:len(raw_choices)]
        q.choices = [
            {'text': text, 'explanation': expl}
            for text, expl in zip(raw_choices, all_explanations)
        ]

        q.points = data['points'] if data['points'] is not None else 1.0

        q.category = str(data['category'] or '').strip()

        raw_level = str(data['level'] or 'easy').strip()
        q.level = _LEVEL_PARSE.get(raw_level.lower(), raw_level.lower())

        q.explanation = str(data['explanation'] or '')
        q.answer_display = str(data.get('answer_display') or '').strip()
        q.shuffle = str(data['shuffle'] or '').strip().lower() in TRUTHY

        raw_strategy = str(data['ma_strategy'] or '').strip()
        q.ma_strategy = _MA_STRATEGY_PARSE.get(
            raw_strategy.lower(), raw_strategy or grading.MA_ALL_OR_NOTHING)

        if q.type in (grading.MC, grading.MA, grading.TF, grading.SA):
            q.correct, errs = parse_correct_spec(
                q.type, data['correct'], len(q.choices))
            q.errors.extend(errs)

        validate_question(q)
    return questions


def write(questions):
    """list[ParsedQuestion] → BytesIO of an importable workbook."""
    from openpyxl import Workbook
    wb = Workbook()
    ws = wb.active
    ws.title = 'Questions'
    _style_sheet(ws)
    ws.append(list(HEADER_LABELS))
    _apply_header_style(ws)

    for q in questions:
        correct_text = (_sa_correct_to_cell(q.correct)
                        if q.type == grading.SA
                        else correct_to_spec(q.type, q.correct))
        raw_choices = q.choices or []
        if raw_choices and isinstance(raw_choices[0], dict):
            choice_texts = [c.get('text', '') for c in raw_choices]
            expls = [c.get('explanation', '') or None for c in raw_choices]
        else:
            choice_texts = [str(c) for c in raw_choices]
            expls = [None] * len(choice_texts)
        choices = choice_texts + [None] * (MAX_CHOICES - len(choice_texts))
        expls = expls + [None] * (MAX_CHOICES - len(expls))
        interleaved = []
        for i in range(MAX_CHOICES):
            interleaved += [choices[i] or None, expls[i] or None]
        ws.append([
            q.code,
            TYPE_DISPLAY.get(q.type, q.type), q.title, q.content,
            *interleaved,
            correct_text,
            q.points, q.category, LEVEL_DISPLAY.get(q.level, q.level),
            q.explanation, 'Yes' if q.shuffle else None,
            MA_STRATEGY_DISPLAY.get(q.ma_strategy, None) if q.type == grading.MA else None,
            getattr(q, 'answer_display', None) if q.type == grading.SA else None,
        ])

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def template():
    """Downloadable starter workbook with example rows and Excel dropdowns."""
    from openpyxl import Workbook
    wb = Workbook()
    ws = wb.active
    ws.title = 'Questions'
    _style_sheet(ws)
    ws.append(list(HEADER_LABELS))
    _apply_header_style(ws)
    for row in EXAMPLE_ROWS:
        ws.append(list(row))

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


# ── Styling helpers ───────────────────────────────────────────────────────────

def _style_sheet(ws):
    """Column widths, freeze pane, wrap text, and dropdown validations."""
    from openpyxl.worksheet.datavalidation import DataValidation

    for col_letter, width in COL_WIDTHS.items():
        ws.column_dimensions[col_letter].width = width

    ws.freeze_panes = 'A2'

    END = 1001

    def _add_dv(formula, col):
        dv = DataValidation(
            type='list', formula1=formula,
            allow_blank=True, showDropDown=False,
            showErrorMessage=True,
        )
        ws.add_data_validation(dv)
        dv.sqref = f'{col}2:{col}{END}'

    _add_dv(f'"{",".join(TYPE_DISPLAY.values())}"',        _col_letter('type'))
    _add_dv(f'"{",".join(LEVEL_DISPLAY.values())}"',       _col_letter('level'))
    _add_dv('"Yes,No"',                                    _col_letter('shuffle'))
    _add_dv(f'"{",".join(MA_STRATEGY_DISPLAY.values())}"', _col_letter('ma_strategy'))


def _apply_header_style(ws):
    """Bold + dark green header row with explanatory cell notes."""
    from openpyxl.comments import Comment
    from openpyxl.styles import Alignment, Font, PatternFill

    hdr_font = Font(bold=True, color='FFFFFF', size=11)
    hdr_fill = PatternFill(start_color='1B5E20', end_color='1B5E20',
                           fill_type='solid')
    hdr_align = Alignment(horizontal='center', vertical='center',
                          wrap_text=True)

    for col_idx, label in enumerate(HEADER_LABELS, start=1):
        cell = ws.cell(row=1, column=col_idx)
        cell.value = label
        cell.font = hdr_font
        cell.fill = hdr_fill
        cell.alignment = hdr_align
        if label in HEADER_NOTES:
            cell.comment = Comment(HEADER_NOTES[label], 'Quiz Import')

    ws.row_dimensions[1].height = 30
